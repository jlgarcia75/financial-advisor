#!/usr/bin/env python3
"""Import an Empower/Pershing "Unrealized Gain Loss" export into a normalized
per-lot cost-basis CSV — the cost basis the statement pipeline otherwise lacks.

The raw export is an .xlsx (sometimes saved with a .csv extension by the browser).
Above the table sits a preamble with the account, as-of date, and summary totals;
the table has one row per tax lot plus security-level rollup rows. Only rows that
carry a Taxlot ID are real lots — rollups are skipped so gains aren't double-counted.
The lot-level Gain/Loss is checksummed against the preamble's "Net Unrealized
Gain/loss"; a mismatch aborts the write.

openpyxl is imported lazily so this module (and its pure normalization logic) loads
without it; only reading a real .xlsx needs it.
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _finance_common import write_csv  # noqa: E402

try:
    import openpyxl
except ImportError:  # only needed to read the .xlsx; pure functions don't require it
    openpyxl = None

VAULT = Path("/Users/jesusgarcia/ObsidianVaults/second-brain/91_finance")
DEFAULT_RAW_DIR = VAULT / "cost_basis"
DEFAULT_OUT_DIR = VAULT / "Reviews/inputs"

HEADER_KEY = "Asset Category"  # first cell of the column-header row
NET_KEY = "Net Unrealized Gain/loss"

# Source column -> normalized column. Extra source columns are ignored.
COLMAP = {
    "Asset Category": "asset_category",
    "Security Description": "security_description",
    "Security Type": "security_type",
    "Quantity": "quantity",
    "Unit Cost": "unit_cost",
    "Current Total Cost": "cost_basis",
    "Market Value": "market_value",
    "Gain/Loss": "gain_loss",
    "Gain/Loss %": "gain_loss_pct",
    "Term": "term",
    "Trade Date": "trade_date",
    "Taxlot ID": "taxlot_id",
    "Covered/NonCovered": "covered",
}
NUMERIC = {"quantity", "unit_cost", "cost_basis", "market_value", "gain_loss", "gain_loss_pct"}
OUT_COLUMNS = ["account_id", "as_of_date", "asset_category", "symbol", "security_description",
               "security_type", "quantity", "unit_cost", "cost_basis", "market_value",
               "gain_loss", "gain_loss_pct", "term", "trade_date", "taxlot_id", "covered"]


def num(x) -> float | None:
    if x is None or x in ("", "-"):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    try:
        return float(str(x).replace(",", "").replace("$", "").strip())
    except ValueError:
        return None


def iso_date(x) -> str:
    if x is None or x in ("", "-", "Multiple"):
        return ""
    if isinstance(x, dt.datetime):
        return x.date().isoformat()
    if isinstance(x, dt.date):
        return x.isoformat()
    s = str(x).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%b %d, %Y"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # e.g. "Jul 30, 2026 8:14 PM EDT" -> pull the "Mon DD, YYYY" portion.
    m = re.search(r"[A-Z][a-z]{2,9}\s+\d{1,2},\s+\d{4}", s)
    if m:
        try:
            return dt.datetime.strptime(m.group(0), "%b %d, %Y").date().isoformat()
        except ValueError:
            pass
    return ""


def read_rows(path: Path) -> list[tuple]:
    """All rows of the first sheet. Reads via a bytes buffer so an .xlsx saved with
    a .csv extension still loads (openpyxl otherwise rejects it on the extension)."""
    if openpyxl is None:
        raise SystemExit("openpyxl is required to read the .xlsx export (pip install openpyxl).")
    wb = openpyxl.load_workbook(io.BytesIO(path.read_bytes()), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return list(ws.iter_rows(values_only=True))


def split_sections(rows: list[tuple]) -> tuple[dict, list[str], list[tuple]]:
    """(preamble metadata, header, data rows) split on the 'Asset Category' header."""
    hdr_idx = next((i for i, r in enumerate(rows)
                    if r and str(r[0]).strip() == HEADER_KEY), None)
    if hdr_idx is None:
        raise SystemExit(f"No '{HEADER_KEY}' header row found — is this an Unrealized Gain Loss export?")
    meta: dict[str, str] = {}
    for r in rows[:hdr_idx]:
        cell = next((c for c in r if c not in (None, "")), None)
        if isinstance(cell, str) and ":" in cell:
            k, _, v = cell.partition(":")
            meta[k.strip()] = v.strip()
    header = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
    return meta, header, rows[hdr_idx + 1:]


def build_lots(meta: dict, header: list[str], data: list[tuple], account_id: str) -> list[dict]:
    """Normalized per-lot rows (Taxlot-ID rows only). Symbol prefers the ticker
    'Symbol' column, falling back to 'Security Identifier'."""
    col = {name: i for i, name in enumerate(header)}
    as_of = iso_date(meta.get("As Of", "")) or iso_date(meta.get("As of", ""))

    def cell(row, name):
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    lots = []
    for row in data:
        if not any(c not in (None, "") for c in row):
            continue
        taxlot = str(cell(row, "Taxlot ID") or "").strip()
        if not taxlot:  # rollup / summary row — skip to avoid double-counting
            continue
        rec = {"account_id": account_id, "as_of_date": as_of}
        for src, dst in COLMAP.items():
            val = cell(row, src)
            if dst in NUMERIC:
                n = num(val)
                rec[dst] = round(n, 2) if n is not None else ""
            elif dst == "trade_date":
                rec[dst] = iso_date(val)
            elif dst == "term":
                rec[dst] = str(val or "").strip().lower()
            else:
                rec[dst] = str(val or "").strip()
        rec["symbol"] = (str(cell(row, "Symbol") or "").strip()
                         or str(cell(row, "Security Identifier") or "").strip())
        lots.append(rec)
    return lots


def verify(meta: dict, lots: list[dict]) -> None:
    stated = num(meta.get(NET_KEY))
    if stated is None:
        return
    total = round(sum(r["gain_loss"] for r in lots if isinstance(r.get("gain_loss"), (int, float))), 2)
    if abs(total - stated) > 1.0:
        raise SystemExit(f"Checksum failed: lot Gain/Loss sums to {total:,.2f} but the report's "
                         f"Net Unrealized is {stated:,.2f}. Not writing.")


def summary(meta: dict, lots: list[dict]) -> str:
    gl = [r["gain_loss"] for r in lots if isinstance(r.get("gain_loss"), (int, float))]
    losses = sum(x for x in gl if x < 0)
    return (f"account {meta.get('Account', '?')} as of {iso_date(meta.get('As Of', ''))}: "
            f"{len(lots)} lots, net unrealized {sum(gl):,.2f}, harvestable losses {losses:,.2f}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import an Empower Unrealized Gain Loss export to a normalized cost-basis CSV.")
    parser.add_argument("export", nargs="?", type=Path, help="Raw .xlsx export (default: newest in cost_basis/).")
    parser.add_argument("--raw-dir", type=Path, default=DEFAULT_RAW_DIR)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--account", help="Override account id (default: from the export preamble).")
    parser.add_argument("--dry-run", action="store_true", help="Print the summary; do not write.")
    args = parser.parse_args()

    src = args.export
    if src is None:
        candidates = sorted(args.raw_dir.glob("*.csv")) + sorted(args.raw_dir.glob("*.xlsx"))
        if not candidates:
            print(f"No export found in {args.raw_dir}.", file=sys.stderr)
            return 2
        src = candidates[-1]
    if not src.exists():
        print(f"Missing export: {src}", file=sys.stderr)
        return 2

    meta, header, data = split_sections(read_rows(src))
    account_id = args.account or meta.get("Account") or src.stem.split("_")[0]
    lots = build_lots(meta, header, data, account_id)
    if not lots:
        print("No tax-lot rows parsed.", file=sys.stderr)
        return 1
    verify(meta, lots)
    print(summary(meta, lots))

    if args.dry_run:
        print(f"(dry run — {len(lots)} lots not written)")
        return 0

    out = args.out_dir / f"cost_basis_{account_id}.csv"
    write_csv(out, lots, preferred=OUT_COLUMNS)
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
